from openpyxl import load_workbook
import sys
import os
import django

SITE_ROOT = os.path.dirname(os.path.realpath(__file__))
print(SITE_ROOT)

sys.path.append(SITE_ROOT)
#sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ['DJANGO_SETTINGS_MODULE'] = 'emile.settings'
django.setup()



from web.models import WallMessages, CourseSectionStudents, CourseSections, CourseSectionStudentsStatus, Courses, Institution, Destinations, Permissions, Users, Teachers,Program, Students, CourseType
from django.db.models import Q

print("Apagando dados anteriores.... ")
WallMessages.objects.all().delete()
CourseSectionStudents.objects.all().delete()
CourseSections.objects.all().delete()
CourseSectionStudentsStatus.objects.all().delete()
Courses.objects.all().delete()
CourseType.objects.all().delete()
Students.objects.all().delete()
Program.objects.all().delete()
Teachers.objects.all().delete()
Users.objects.all().delete()
Permissions.objects.all().delete()
Institution.objects.all().delete()
Destinations.objects.all().delete()

print("Incluindo Institutions ....")
wb = load_workbook('web/planilhas/Institutions.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=5, max_row=2):
    institution = Institution()
    for cell in row:
        mylist.append(cell.value)
    institution.name = mylist[0].strip()
    institution.abbreviation = mylist[1].strip()
    institution.cnpj = mylist[2].strip()
    institution.address = mylist[3].strip()
    institution.current_program_section = mylist[4]
    institution.save()
    mylist.clear()
#print(institution)

print("Incluindo Destinations ....")
wb = load_workbook('web/planilhas/Destinations.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=2, max_row=5):
    destinations = Destinations()
    for cell in row:
        mylist.append(cell.value)
    destinations.name = mylist[0].strip()
    destinations.url_service = mylist[1].strip()
    destinations.save()
    mylist.clear()
#print(destinations)


print("Incluindo Permissions ....")
permission = Permissions()
permission.description = "student"
permission.save()

permission = Permissions()
permission.description = "coordinator"
permission.save()
permission.messagedestinations = Destinations.objects.all()

permission = Permissions()
permission.description = "teacher"
permission.save()
permission.messagedestinations = Destinations.objects.filter(Q(url_service="sendMessagesToStudentsOfATeacher")|Q(url_service="sendMessageToStudentsOfACourseSection"))

print("Incluindo Teachers ....")
wb = load_workbook('web/planilhas/Teachers.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=5, max_row=190):
    teacher = Teachers()
    user = Users()
    for cell in row:
        mylist.append(cell.value)
    teacher.siape = str(mylist[0]).strip()
    user.name = mylist[1].strip()
    user.email = mylist[2].strip()
    user.password = mylist[3].strip()
    user.permission = Permissions.objects.get(description="teacher")
    user.save()
    teacher.user = user
    teacher.save()
    mylist.clear()
#print(mylist)


print("Incluindo Programs ....")
wb = load_workbook('web/planilhas/Programs.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=3, max_row=12):
    teacher = Teachers.objects.get(siape="1625546")# como não vem coloquei Renato Novais
    program = Program()
    for cell in row:
        mylist.append(cell.value)
    program.id = mylist[0]
    program.name = mylist[1].strip()
    program.abbreviation = mylist[2].strip()
    program.coordinator = teacher
    program.institution = Institution.objects.get(abbreviation="IFBA-SSA")
    program.save()
    mylist.clear()
#print(mylist) 1625546


print("Incluindo Students ....")
wb = load_workbook('web/planilhas/Students.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=6, max_row=1937):
    student = Students()
    user = Users()
    for cell in row:
        mylist.append(cell.value)
    student.enrollment = mylist[2].strip()
    student.program = Program.objects.get(id=mylist[0])
    user.name = mylist[3].strip()
    user.email = mylist[4].strip()
    user.password = mylist[5].strip()
    user.permission = Permissions.objects.get(description="student")
    user.save()
    student.user = user
    student.save()
    mylist.clear()

print("Incluindo CourseType ....")
coursetype = CourseType()
coursetype.description = "Optativa"
coursetype.save()

coursetype = CourseType()
coursetype.description = "Obrigatória"
coursetype.save()


print("Incluindo CourseSectionStudentsStatus ....")
coursesectionstudentsstatus = CourseSectionStudentsStatus()
coursesectionstudentsstatus.description = "Trancado"
coursesectionstudentsstatus.save()

coursesectionstudentsstatus = CourseSectionStudentsStatus()
coursesectionstudentsstatus.description = "Cursando"
coursesectionstudentsstatus.save()


print("Incluindo Courses ....")
wb = load_workbook('web/planilhas/Courses.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=7, max_row=1042):
    course = Courses()
    for cell in row:
        mylist.append(cell.value)
    course.code = mylist[3].strip()
    course.name = mylist[4].strip()
    course.credits =  mylist[5]
    course.hours = mylist[6]
    course.program_section = mylist[2]
    course.course_type = CourseType.objects.get(description="Obrigatória")
    course.program = Program.objects.get(id=mylist[0])
    course.save()
    mylist.clear()

print("Incluindo CourseSections ....")
wb = load_workbook('web/planilhas/CourseSections.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=5, max_row=480):
    coursesection = CourseSections()
    for cell in row:
        mylist.append(cell.value)
    coursesection.code = mylist[2].strip()
    coursesection.name = ('{}-{}-{}'.format(mylist[3].strip(), mylist[4].strip(),mylist[2].strip()))
    listeofcodes =Courses.objects.filter(code=mylist[3].strip())
    coursesection.course =  listeofcodes[0]
    selectedteacher = Teachers.objects.get(siape=mylist[0].strip())
    coursesection.teacher = selectedteacher
    coursesection.course_section_period = "1"
    coursesection.save()
    mylist.clear()


print("Incluindo CourseSectionsStudends ....")
wb = load_workbook('web/planilhas/CourseSectionsStudends.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sheet1')
#mylist = []
for row in ws.get_squared_range(min_col=1, min_row=2, max_col=5, max_row=9382):
    coursesectionstudent = CourseSectionStudents()
    for cell in row:
        mylist.append(cell.value)
    selectedstudent = Students.objects.get(enrollment=mylist[3].strip())
    coursesectionstudent.student = selectedstudent
    names = '{}-{}-{}'.format(mylist[1].strip(),mylist[2].strip(),mylist[0].strip())
    #print("Name: >%s<" % names)
    selectedcoursesection = CourseSections.objects.get(name=names)
    print(selectedcoursesection)
    coursesectionstudent.course_section = selectedcoursesection
    coursesectionstudent.status = CourseSectionStudentsStatus.objects.get(description="Cursando")
    coursesectionstudent.save()
    mylist.clear()
